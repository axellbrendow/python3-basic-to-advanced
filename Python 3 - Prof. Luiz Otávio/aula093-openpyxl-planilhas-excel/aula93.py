import openpyxl

print('Carregando arquivo...')
pedidos = openpyxl.load_workbook('pedidos.xlsx')

print('Processando os dados...')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']

print('\nValor da célula b4')
print(planilha1['b4'].value)

print('\nValores da coluna b')
for celula in planilha1['b']:
    print(celula.value)

print('\nRange a1:c2')
for linha in planilha1['a1:c2']:
    for coluna in linha:
        print(f'{str(coluna.value):15}', end='')

    print()

print('\nToda a planilha')
for linha in planilha1:
    for coluna in linha:
        print(f'{str(coluna.value):15}', end='')

    print()

from random import uniform

print('\nGerando uma nova planilha')
planilha1['b3'].value = 2200

for linha in range(5, 16):
    num_pedido = linha - 1
    planilha1.cell(linha, 1).value = num_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')

print('\nCriando planilha do zero')
planilhas = openpyxl.Workbook()

planilhas.create_sheet('Planilha1', 0)
planilhas.create_sheet('Planilha2', 1)

planilha1 = planilhas['Planilha1']
planilha2 = planilhas['Planilha2']

for linha in range(1, 16):
    planilha1.cell(linha, 1).value = 'p1c1: ' + str(round(uniform(10, 100), 2))
    planilha1.cell(linha, 2).value = 'p1c2: ' + str(round(uniform(10, 100), 2))
    planilha1.cell(linha, 3).value = 'p1c3: ' + str(round(uniform(10, 100), 2))

for linha in range(1, 16):
    planilha2.cell(linha, 1).value = 'p2c1: ' + str(round(uniform(10, 100), 2))
    planilha2.cell(linha, 2).value = 'p2c2: ' + str(round(uniform(10, 100), 2))
    planilha2.cell(linha, 3).value = 'p2c3: ' + str(round(uniform(10, 100), 2))

planilhas.save('nova_planilha2.xlsx')
